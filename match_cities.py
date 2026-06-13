from __future__ import annotations

import argparse
import csv
import itertools
import re
import shutil
import sys
import threading
import time
import unicodedata
from contextlib import contextmanager
from dataclasses import dataclass
from pathlib import Path
from typing import Callable, Iterable, Iterator

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


RC_SHEET = "rc"
SHREP_SHEET = "shrep"

BRAND_NAME = "Andy's LineMatcher"
TAGLINE = "Smart city matching for RC <-> Shipment Report"

CONSOLE_BANNER_LINES = [
    "  andy",
    "",
    "  ██╗     ██╗███╗   ██╗███████╗",
    "  ██║     ██║████╗  ██║██╔════╝",
    "  ██║     ██║██╔██╗ ██║█████╗",
    "  ██║     ██║██║╚██╗██║██╔══╝",
    "  ███████╗██║██║ ╚████║███████╗",
    "  ╚══════╝╚═╝╚═╝  ╚═══╝╚══════╝",
    "",
    "  ███╗   ███╗ █████╗ ████████╗ ██████╗██╗  ██╗███████╗██████╗",
    "  ████╗ ████║██╔══██╗╚══██╔══╝██╔════╝██║  ██║██╔════╝██╔══██╗",
    "  ██╔████╔██║███████║   ██║   ██║     ███████║█████╗  ██████╔╝",
    "  ██║╚██╔╝██║██╔══██║   ██║   ██║     ██╔══██║██╔══╝  ██╔══██╗",
    "  ██║ ╚═╝ ██║██║  ██║   ██║   ╚██████╗██║  ██║███████╗██║  ██║",
    "  ╚═╝     ╚═╝╚═╝  ╚═╝   ╚═╝    ╚═════╝╚═╝  ╚═╝╚══════╝╚═╝  ╚═╝",
]

MINI_BANNER_LINES = [
    "  andy",
    "  LINE MATCHER",
]

RC_CITY_COL = "Destination City"
RC_COUNTRY_COL = "Destination Country"
SHREP_CITY_COL = "Delivery City"
SHREP_COUNTRY_COL = "Arrival Country"

ALIAS_HEADERS = [
    "shrep_city",
    "arrival_country",
    "rc_city",
    "destination_country",
]

GENERIC_CITY_WORDS = {
    "CITY",
    "CTY",
    "TOWN",
    "DISTRICT",
    "PROVINCE",
    "GOVERNORATE",
    "MUNICIPALITY",
}

COUNTRY_ALIASES = {
    "AE": {"AE", "UAE", "UNITED ARAB EMIRATES"},
    "AU": {"AU", "AUS", "AUSTRALIA"},
    "BH": {"BH", "BAHRAIN"},
    "IL": {"IL", "ISRAEL"},
    "KW": {"KW", "KUWAIT"},
    "OM": {"OM", "OMAN"},
}

ANSI_COLORS = {
    "reset": "\033[0m",
    "bold": "\033[1m",
    "dim": "\033[2m",
    "ember": "\033[38;2;200;104;68m",
    "pistachio": "\033[38;2;155;188;92m",
    "cyan": "\033[36m",
    "green": "\033[38;2;155;188;92m",
    "yellow": "\033[33m",
    "red": "\033[31m",
    "white": "\033[97m",
}


def enable_utf8_console() -> None:
    if sys.platform != "win32":
        return

    try:
        import ctypes

        kernel32 = ctypes.windll.kernel32
        kernel32.SetConsoleOutputCP(65001)
        kernel32.SetConsoleCP(65001)
    except Exception:
        pass

    for stream_name in ("stdout", "stderr"):
        stream = getattr(sys, stream_name, None)
        reconfigure = getattr(stream, "reconfigure", None)
        if callable(reconfigure):
            try:
                reconfigure(encoding="utf-8", errors="replace")
            except Exception:
                pass


def color_enabled() -> bool:
    return sys.stdout.isatty() and not bool(os_environ_no_color())


def os_environ_no_color() -> str | None:
    import os

    return os.environ.get("NO_COLOR")


def colorize(text: object, color: str, enabled: bool | None = None) -> str:
    if enabled is None:
        enabled = color_enabled()
    value = str(text)
    if not enabled:
        return value
    prefix = ANSI_COLORS.get(color, "")
    if not prefix:
        return value
    return f"{prefix}{value}{ANSI_COLORS['reset']}"


@dataclass(frozen=True)
class Candidate:
    rc_row_id: int
    destination_country: str
    destination_city: str
    normalized_city: str
    variant: str


@dataclass(frozen=True)
class MatchResult:
    status: str
    method: str
    score: float
    rc_row_id: int | None
    destination_country: str | None
    destination_city: str | None
    normalized_delivery_city: str
    candidate_count: int
    top_candidates: list[tuple[str, str, float]]


class Spinner:
    def __init__(self, label: str, enabled: bool = True) -> None:
        self.label = label
        self.enabled = enabled and sys.stdout.isatty()
        self._stop = threading.Event()
        self._thread: threading.Thread | None = None

    def __enter__(self) -> Spinner:
        if self.enabled:
            self._thread = threading.Thread(target=self._spin, daemon=True)
            self._thread.start()
        else:
            print(f"[..] {self.label}")
        return self

    def __exit__(self, exc_type, exc, tb) -> None:  # type: ignore[no-untyped-def]
        if self.enabled:
            self._stop.set()
            if self._thread:
                self._thread.join()
            sys.stdout.write("\r" + " " * (len(self.label) + 12) + "\r")
        if exc_type is None:
            print(f"{colorize('[ok]', 'green')} {self.label}")
        else:
            print(f"{colorize('[!!]', 'red')} {self.label}")

    def _spin(self) -> None:
        frames = itertools.cycle(("|", "/", "-", "\\"))
        while not self._stop.is_set():
            sys.stdout.write(f"\r{colorize(f'[{next(frames)}]', 'cyan')} {self.label}")
            sys.stdout.flush()
            time.sleep(0.08)


@contextmanager
def step(label: str) -> Iterator[None]:
    with Spinner(label):
        yield


def compact_header(header: object) -> str:
    return re.sub(r"\s+", " ", str(header or "")).strip()


def normalize_country(value: object) -> str:
    return normalize_basic(value).replace(" ", "")


def normalize_basic(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip().upper()
    text = unicodedata.normalize("NFKD", text)
    text = text.encode("ascii", "ignore").decode("ascii")
    text = re.sub(r"[^A-Z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def normalize_city(value: object, country: object = "") -> str:
    text = normalize_basic(value)
    if not text:
        return ""

    tokens = text.split()
    country_tokens = COUNTRY_ALIASES.get(normalize_country(country), set())

    if len(tokens) > 1:
        tokens = [token for token in tokens if token not in GENERIC_CITY_WORDS]

    if len(tokens) > 1 and tokens[-1] in country_tokens:
        tokens = tokens[:-1]

    return " ".join(tokens)


def city_variants(city: object, country: object) -> set[str]:
    base = normalize_city(city, country)
    variants = {base} if base else set()

    for part in re.split(r"[;,/|]", str(city or "")):
        normalized = normalize_city(part, country)
        if normalized:
            variants.add(normalized)

    return variants


def sequence_score(left: str, right: str) -> float:
    from difflib import SequenceMatcher

    return 100.0 * SequenceMatcher(None, left, right).ratio()


def token_sort_score(left: str, right: str) -> float:
    sorted_left = " ".join(sorted(left.split()))
    sorted_right = " ".join(sorted(right.split()))
    return sequence_score(sorted_left, sorted_right)


def subset_score(left: str, right: str) -> float:
    left_tokens = set(left.split())
    right_tokens = set(right.split())
    if not left_tokens or not right_tokens:
        return 0.0

    shorter = left_tokens if len(left_tokens) <= len(right_tokens) else right_tokens
    longer = right_tokens if shorter is left_tokens else left_tokens
    shorter_text = " ".join(sorted(shorter))
    if len(shorter_text) < 4:
        return 0.0

    if shorter.issubset(longer):
        token_gap = len(longer) - len(shorter)
        return max(90.0, 97.0 - token_gap * 2.0)

    return 0.0


def city_score(left: str, right: str) -> float:
    if not left or not right:
        return 0.0
    if left == right:
        return 100.0
    return max(
        sequence_score(left, right),
        token_sort_score(left, right),
        subset_score(left, right),
    )


def read_sheet_records(path: Path, sheet_name: str) -> list[dict[str, object]]:
    workbook = load_workbook(path, read_only=True, data_only=False)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Sheet not found: {sheet_name}")

    sheet = workbook[sheet_name]
    rows = sheet.iter_rows(values_only=True)
    try:
        headers = [compact_header(value) for value in next(rows)]
    except StopIteration as exc:
        raise ValueError(f"Sheet is empty: {sheet_name}") from exc

    records = []
    for row in rows:
        if all(value is None for value in row):
            continue
        records.append({headers[index]: value for index, value in enumerate(row)})
    workbook.close()
    return records


def load_input(path: Path) -> tuple[list[dict[str, object]], list[dict[str, object]]]:
    rc = read_sheet_records(path, RC_SHEET)
    shrep = read_sheet_records(path, SHREP_SHEET)

    missing_rc = {RC_CITY_COL, RC_COUNTRY_COL} - set(rc[0].keys() if rc else [])
    missing_shrep = {SHREP_CITY_COL, SHREP_COUNTRY_COL} - set(
        shrep[0].keys() if shrep else []
    )
    if missing_rc or missing_shrep:
        raise ValueError(
            "Missing required columns: "
            f"rc={sorted(missing_rc)}, shrep={sorted(missing_shrep)}"
        )

    for index, row in enumerate(rc, start=1):
        row["rc_row_id"] = index
    for index, row in enumerate(shrep, start=1):
        row["shrep_row_id"] = index

    return rc, shrep


def build_candidates(rc: list[dict[str, object]]) -> list[Candidate]:
    candidates: list[Candidate] = []
    for row in rc:
        country = normalize_country(row[RC_COUNTRY_COL])
        city = str(row[RC_CITY_COL]).strip()
        for variant in city_variants(city, country):
            candidates.append(
                Candidate(
                    rc_row_id=int(row["rc_row_id"]),
                    destination_country=country,
                    destination_city=city,
                    normalized_city=variant,
                    variant=variant,
                )
            )
    return candidates


def load_aliases(path: Path | None) -> dict[tuple[str, str], tuple[str, str]]:
    if path is None or not path.exists():
        return {}

    aliases: dict[tuple[str, str], tuple[str, str]] = {}
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            source_country = normalize_country(row.get("arrival_country", ""))
            source_city = normalize_city(row.get("shrep_city", ""), source_country)
            target_country = normalize_country(
                row.get("destination_country", source_country)
            )
            target_city = normalize_city(row.get("rc_city", ""), target_country)
            if source_city and target_city:
                aliases[(source_country, source_city)] = (target_country, target_city)
    return aliases


def dedupe_ranked(
    ranked: Iterable[tuple[Candidate, float]],
) -> list[tuple[Candidate, float]]:
    best_by_row: dict[int, tuple[Candidate, float]] = {}
    for candidate, score in ranked:
        current = best_by_row.get(candidate.rc_row_id)
        if current is None or score > current[1]:
            best_by_row[candidate.rc_row_id] = (candidate, score)
    return sorted(best_by_row.values(), key=lambda item: item[1], reverse=True)


def match_one(
    delivery_city: object,
    arrival_country: object,
    candidates: list[Candidate],
    aliases: dict[tuple[str, str], tuple[str, str]],
    auto_threshold: float,
    review_threshold: float,
    min_margin: float,
    allow_cross_country: bool,
) -> MatchResult:
    country = normalize_country(arrival_country)
    normalized_city = normalize_city(delivery_city, country)
    same_country = [c for c in candidates if c.destination_country == country]
    candidate_pool = same_country or (candidates if allow_cross_country else [])

    alias_target = aliases.get((country, normalized_city))
    if alias_target:
        target_country, target_city = alias_target
        alias_candidates = [
            c
            for c in candidates
            if c.destination_country == target_country and c.normalized_city == target_city
        ]
        if alias_candidates:
            candidate = alias_candidates[0]
            return MatchResult(
                status="auto_matched",
                method="alias",
                score=100.0,
                rc_row_id=candidate.rc_row_id,
                destination_country=candidate.destination_country,
                destination_city=candidate.destination_city,
                normalized_delivery_city=normalized_city,
                candidate_count=len(candidate_pool),
                top_candidates=[
                    (candidate.destination_country, candidate.destination_city, 100.0)
                ],
            )

    if not normalized_city or not candidate_pool:
        return MatchResult(
            status="unmatched",
            method="none",
            score=0.0,
            rc_row_id=None,
            destination_country=None,
            destination_city=None,
            normalized_delivery_city=normalized_city,
            candidate_count=len(candidate_pool),
            top_candidates=[],
        )

    ranked = dedupe_ranked(
        (candidate, city_score(normalized_city, candidate.normalized_city))
        for candidate in candidate_pool
    )
    top_candidate, top_score = ranked[0]
    second_score = ranked[1][1] if len(ranked) > 1 else 0.0
    margin = top_score - second_score

    if top_score >= auto_threshold and (top_score == 100.0 or margin >= min_margin):
        status = "auto_matched"
    elif top_score >= review_threshold:
        status = "review_needed"
    else:
        status = "unmatched"

    return MatchResult(
        status=status,
        method="fuzzy",
        score=round(top_score, 1),
        rc_row_id=top_candidate.rc_row_id if status != "unmatched" else None,
        destination_country=top_candidate.destination_country
        if status != "unmatched"
        else None,
        destination_city=top_candidate.destination_city if status != "unmatched" else None,
        normalized_delivery_city=normalized_city,
        candidate_count=len(candidate_pool),
        top_candidates=[
            (candidate.destination_country, candidate.destination_city, round(score, 1))
            for candidate, score in ranked[:3]
        ],
    )


def result_to_columns(result: MatchResult) -> dict[str, object]:
    columns: dict[str, object] = {
        "match_status": result.status,
        "match_method": result.method,
        "match_score": result.score,
        "matched_rc_row_id": result.rc_row_id,
        "matched_destination_country": result.destination_country,
        "matched_destination_city": result.destination_city,
        "normalized_delivery_city": result.normalized_delivery_city,
        "candidate_count": result.candidate_count,
    }

    for index in range(3):
        prefix = f"candidate_{index + 1}"
        if index < len(result.top_candidates):
            country, city, score = result.top_candidates[index]
            columns[f"{prefix}_country"] = country
            columns[f"{prefix}_city"] = city
            columns[f"{prefix}_score"] = score
        else:
            columns[f"{prefix}_country"] = None
            columns[f"{prefix}_city"] = None
            columns[f"{prefix}_score"] = None
    return columns


def apply_manual_choice(record: dict[str, object], choice: int) -> dict[str, object]:
    candidate_country = record.get(f"candidate_{choice}_country")
    candidate_city = record.get(f"candidate_{choice}_city")
    candidate_score = record.get(f"candidate_{choice}_score")
    if not candidate_country or not candidate_city:
        raise ValueError(f"Candidate {choice} is not available")

    updated = dict(record)
    updated["match_status"] = "manual_matched"
    updated["match_method"] = "manual"
    updated["match_score"] = candidate_score
    updated["matched_destination_country"] = candidate_country
    updated["matched_destination_city"] = candidate_city
    return updated


def alias_row_from_match(record: dict[str, object]) -> dict[str, object]:
    return {
        "shrep_city": record.get(SHREP_CITY_COL),
        "arrival_country": record.get(SHREP_COUNTRY_COL),
        "rc_city": record.get("matched_destination_city"),
        "destination_country": record.get("matched_destination_country"),
    }


def unique_records(
    records: Iterable[dict[str, object]],
    keys: tuple[str, ...],
) -> list[dict[str, object]]:
    seen = set()
    output = []
    for record in records:
        key = tuple(record.get(name) for name in keys)
        if key in seen:
            continue
        seen.add(key)
        output.append(record)
    return output


def build_outputs(
    rc: list[dict[str, object]],
    shrep: list[dict[str, object]],
    candidates: list[Candidate],
    aliases: dict[tuple[str, str], tuple[str, str]],
    auto_threshold: float,
    review_threshold: float,
    min_margin: float,
    allow_cross_country: bool,
) -> dict[str, list[dict[str, object]]]:
    rc_by_id = {int(row["rc_row_id"]): row for row in rc}
    matched_rows = []
    for row in shrep:
        result = match_one(
            row[SHREP_CITY_COL],
            row[SHREP_COUNTRY_COL],
            candidates,
            aliases,
            auto_threshold,
            review_threshold,
            min_margin,
            allow_cross_country,
        )
        combined = {**row, **result_to_columns(result)}
        rc_id = combined.get("matched_rc_row_id")
        if rc_id:
            combined.update({f"rc.{key}": value for key, value in rc_by_id[int(rc_id)].items()})
        matched_rows.append(combined)

    city_cols = [
        SHREP_COUNTRY_COL,
        SHREP_CITY_COL,
        "normalized_delivery_city",
        "match_status",
        "match_method",
        "match_score",
        "matched_destination_country",
        "matched_destination_city",
        "matched_rc_row_id",
        "candidate_1_country",
        "candidate_1_city",
        "candidate_1_score",
        "candidate_2_country",
        "candidate_2_city",
        "candidate_2_score",
        "candidate_3_country",
        "candidate_3_city",
        "candidate_3_score",
    ]
    city_map = [
        {column: row.get(column) for column in city_cols}
        for row in unique_records(
            matched_rows,
            (
                SHREP_COUNTRY_COL,
                SHREP_CITY_COL,
                "normalized_delivery_city",
                "match_status",
                "matched_destination_city",
            ),
        )
    ]
    city_map.sort(
        key=lambda row: (
            str(row.get(SHREP_COUNTRY_COL) or ""),
            str(row.get(SHREP_CITY_COL) or ""),
        )
    )

    review = [
        row
        for row in city_map
        if row.get("match_status") in {"review_needed", "unmatched"}
    ]
    suggested_aliases = [
        alias_row_from_match(row)
        for row in city_map
        if row.get("match_status") in {"auto_matched", "manual_matched"}
    ]

    summary = make_summary(
        shrep_rows=len(shrep),
        rc_rows=len(rc),
        city_map=city_map,
        auto_threshold=auto_threshold,
        review_threshold=review_threshold,
        min_margin=min_margin,
        allow_cross_country=allow_cross_country,
    )

    return {
        "summary": summary,
        "city_map": city_map,
        "matched_rows": matched_rows,
        "review_needed": review,
        "suggested_aliases": suggested_aliases,
    }


def make_summary(
    shrep_rows: int,
    rc_rows: int,
    city_map: list[dict[str, object]],
    auto_threshold: float,
    review_threshold: float,
    min_margin: float,
    allow_cross_country: bool,
) -> list[dict[str, object]]:
    return [
        {"metric": "shrep_rows", "value": shrep_rows},
        {"metric": "rc_rows", "value": rc_rows},
        {"metric": "unique_shrep_city_country_pairs", "value": len(city_map)},
        {
            "metric": "auto_matched_pairs",
            "value": count_status(city_map, "auto_matched"),
        },
        {
            "metric": "manual_matched_pairs",
            "value": count_status(city_map, "manual_matched"),
        },
        {
            "metric": "review_needed_pairs",
            "value": count_status(city_map, "review_needed"),
        },
        {"metric": "unmatched_pairs", "value": count_status(city_map, "unmatched")},
        {"metric": "auto_threshold", "value": auto_threshold},
        {"metric": "review_threshold", "value": review_threshold},
        {"metric": "min_margin", "value": min_margin},
        {"metric": "allow_cross_country", "value": allow_cross_country},
    ]


def count_status(records: list[dict[str, object]], status: str) -> int:
    return sum(1 for record in records if record.get("match_status") == status)


def print_banner() -> None:
    terminal = shutil.get_terminal_size(fallback=(120, 40))
    lines = CONSOLE_BANNER_LINES if terminal.columns >= 76 else MINI_BANNER_LINES
    print()
    for index, line in enumerate(lines):
        tone = "dim" if "andy" in line.lower() else "ember"
        print(colorize(line, tone))
    print(colorize(TAGLINE, "dim"))
    print(colorize("-" * 64, "ember"))


def render_matched_pairs(
    city_map: list[dict[str, object]],
    enabled: bool | None = None,
) -> list[str]:
    matched = [
        row
        for row in city_map
        if row.get("match_status") in {"auto_matched", "manual_matched"}
    ]
    lines: list[str] = []
    for row in matched:
        source = f"{review_value(row.get(SHREP_COUNTRY_COL))} / {review_value(row.get(SHREP_CITY_COL))}"
        target = (
            f"{review_value(row.get('matched_destination_country'))} / "
            f"{review_value(row.get('matched_destination_city'))}"
        )
        method = review_value(row.get("match_method"))
        score = review_value(row.get("match_score"))
        line = f"  {source:<28} -> {target:<34} | {method:<6} | score {score}"
        lines.append(colorize(line, "green", enabled=enabled))
    return lines


def print_summary(
    summary: list[dict[str, object]],
    city_map: list[dict[str, object]],
    output: Path,
) -> None:
    values = {str(row["metric"]): row["value"] for row in summary}
    print()
    print(colorize("Summary", "cyan"))
    print(colorize("-" * 64, "cyan"))
    print(f"  SHREP rows         {values['shrep_rows']}")
    print(f"  RC rows            {values['rc_rows']}")
    print(f"  City pairs         {values['unique_shrep_city_country_pairs']}")
    print(f"  Auto matched       {colorize(values['auto_matched_pairs'], 'green')}")
    print(f"  Manual matched     {colorize(values['manual_matched_pairs'], 'green')}")
    print(f"  Review needed      {colorize(values['review_needed_pairs'], 'yellow')}")
    print(f"  Unmatched          {colorize(values['unmatched_pairs'], 'yellow')}")
    print(
        "  Decision pending   "
        f"{colorize(int(values['review_needed_pairs']) + int(values['unmatched_pairs']), 'yellow')}"
    )
    print(colorize("-" * 64, "cyan"))
    print(colorize("Matched pairs", "cyan"))
    lines = render_matched_pairs(city_map)
    if lines:
        for line in lines:
            print(line)
    else:
        print(colorize("  none", "dim"))
    print(colorize("-" * 64, "cyan"))
    print(f"Output: {output}")


def should_prompt_for_review(
    explicit_review: bool,
    review_rows: list[dict[str, object]],
    no_args_mode: bool,
) -> bool:
    return no_args_mode and not explicit_review and bool(review_rows)


def should_pause_on_exit(no_args_mode: bool, frozen: bool) -> bool:
    return no_args_mode or frozen


def ask_yes_no(prompt: str, default: bool = False) -> bool:
    suffix = "[Y/n]" if default else "[y/N]"
    try:
        answer = input(f"{prompt} {suffix} ").strip().lower()
    except EOFError:
        return default
    if not answer:
        return default
    return answer in {"y", "yes", "t", "tak"}


def pause_console() -> None:
    try:
        input("\nPress Enter to close...")
    except EOFError:
        pass


def review_line(width: int = 64) -> str:
    return colorize("-" * width, "cyan")


def review_title(text: str, width: int = 64) -> str:
    text = f" {text.strip()} "
    return colorize(text.center(width, "-"), "cyan")


def review_value(value: object) -> str:
    if value is None:
        return ""
    return str(value)


def interactive_review(
    city_map: list[dict[str, object]],
    input_func: Callable[[str], str] = input,
    output_func: Callable[[str], None] = print,
) -> list[dict[str, object]]:
    updated = []
    queue = [
        row
        for row in city_map
        if row.get("match_status") in {"review_needed", "unmatched"}
    ]

    if not queue:
        output_func(colorize("No records need manual review.", "green"))
        return city_map

    output_func("")
    output_func(review_title("Manual review"))
    output_func(colorize("Pick 1-3 to approve a candidate.", "white"))
    output_func(
        colorize(
            "Press Enter, q, or s to skip this city. Press x to finish review.",
            "dim",
        )
    )
    output_func(review_line())

    queue_keys = {
        (row.get(SHREP_COUNTRY_COL), row.get(SHREP_CITY_COL), row.get("normalized_delivery_city"))
        for row in queue
    }

    replacements: dict[tuple[object, object, object], dict[str, object]] = {}
    for index, row in enumerate(queue, start=1):
        key = (
            row.get(SHREP_COUNTRY_COL),
            row.get(SHREP_CITY_COL),
            row.get("normalized_delivery_city"),
        )
        output_func("")
        output_func(review_title(f"Decision {index} of {len(queue)}"))
        output_func(f"Country : {review_value(row.get(SHREP_COUNTRY_COL))}")
        output_func(f"City    : {review_value(row.get(SHREP_CITY_COL))}")
        output_func(f"Status  : {colorize(review_value(row.get('match_status')), 'yellow')}")
        output_func("")
        output_func(colorize("Candidates", "cyan"))
        for candidate_index in range(1, 4):
            city = row.get(f"candidate_{candidate_index}_city")
            if not city:
                continue
            country = row.get(f"candidate_{candidate_index}_country")
            score = row.get(f"candidate_{candidate_index}_score")
            output_func(
                f"  {candidate_index}) "
                f"{review_value(country):<3}  "
                f"{review_value(city):<34} "
                f"score {colorize(review_value(score), 'yellow')}"
            )

        output_func("")
        prompt = colorize("Choose [1-3], skip [q/s/Enter], finish [x]: ", "cyan")
        answer = input_func(prompt).strip().lower()
        if answer == "x":
            break
        if answer in {"q", "s", ""}:
            output_func(colorize("  skipped", "yellow"))
            continue
        if answer in {"1", "2", "3"}:
            try:
                replacements[key] = apply_manual_choice(row, int(answer))
                accepted = replacements[key]
                output_func(
                    colorize("  accepted -> ", "green")
                    +
                    f"{accepted.get('matched_destination_country')} / "
                    f"{accepted.get('matched_destination_city')}"
                )
            except ValueError as exc:
                output_func(colorize(f"  skipped: {exc}", "yellow"))
            continue
        output_func(colorize("  skipped: unknown command", "yellow"))

    for row in city_map:
        key = (
            row.get(SHREP_COUNTRY_COL),
            row.get(SHREP_CITY_COL),
            row.get("normalized_delivery_city"),
        )
        updated.append(replacements.get(key, row) if key in queue_keys else row)
    return updated


def apply_city_map_to_rows(
    matched_rows: list[dict[str, object]],
    city_map: list[dict[str, object]],
) -> list[dict[str, object]]:
    by_key = {
        (row.get(SHREP_COUNTRY_COL), row.get(SHREP_CITY_COL), row.get("normalized_delivery_city")): row
        for row in city_map
    }
    updated = []
    for row in matched_rows:
        key = (
            row.get(SHREP_COUNTRY_COL),
            row.get(SHREP_CITY_COL),
            row.get("normalized_delivery_city"),
        )
        city_record = by_key.get(key)
        if city_record:
            merged = dict(row)
            for field in [
                "match_status",
                "match_method",
                "match_score",
                "matched_destination_country",
                "matched_destination_city",
            ]:
                merged[field] = city_record.get(field)
            updated.append(merged)
        else:
            updated.append(row)
    return updated


def refresh_review_and_alias_sheets(
    outputs: dict[str, list[dict[str, object]]],
    shrep_rows: int,
    rc_rows: int,
    auto_threshold: float,
    review_threshold: float,
    min_margin: float,
    allow_cross_country: bool,
) -> None:
    city_map = outputs["city_map"]
    outputs["matched_rows"] = apply_city_map_to_rows(outputs["matched_rows"], city_map)
    outputs["review_needed"] = [
        row
        for row in city_map
        if row.get("match_status") in {"review_needed", "unmatched"}
    ]
    outputs["suggested_aliases"] = [
        alias_row_from_match(row)
        for row in city_map
        if row.get("match_status") in {"auto_matched", "manual_matched"}
    ]
    outputs["summary"] = make_summary(
        shrep_rows=shrep_rows,
        rc_rows=rc_rows,
        city_map=city_map,
        auto_threshold=auto_threshold,
        review_threshold=review_threshold,
        min_margin=min_margin,
        allow_cross_country=allow_cross_country,
    )


def ordered_headers(records: list[dict[str, object]]) -> list[str]:
    headers: list[str] = []
    seen = set()
    for record in records:
        for key in record:
            if key not in seen:
                seen.add(key)
                headers.append(key)
    return headers


def write_sheet(workbook: Workbook, name: str, records: list[dict[str, object]]) -> None:
    sheet = workbook.create_sheet(title=name)
    headers = ordered_headers(records)
    if not headers:
        sheet.append(["empty"])
        return

    sheet.append(headers)
    for record in records:
        sheet.append([record.get(header) for header in headers])

    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column in sheet.columns:
        letter = get_column_letter(column[0].column)
        max_len = max(len(str(cell.value or "")) for cell in column[:200])
        sheet.column_dimensions[letter].width = min(max(max_len + 2, 10), 46)


def write_outputs(outputs: dict[str, list[dict[str, object]]], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.remove(workbook.active)
    for sheet_name, records in outputs.items():
        write_sheet(workbook, sheet_name, records)
    workbook.save(output_path)


def append_aliases(path: Path, rows: list[dict[str, object]]) -> int:
    if not rows:
        return 0

    existing = set()
    if path.exists():
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            for row in reader:
                existing.add(tuple(row.get(header, "") for header in ALIAS_HEADERS))

    new_rows = []
    for row in rows:
        key = tuple(str(row.get(header) or "") for header in ALIAS_HEADERS)
        if all(key) and key not in existing:
            existing.add(key)
            new_rows.append({header: row.get(header) for header in ALIAS_HEADERS})

    path.parent.mkdir(parents=True, exist_ok=True) if path.parent != Path(".") else None
    write_header = not path.exists() or path.stat().st_size == 0
    with path.open("a", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=ALIAS_HEADERS)
        if write_header:
            writer.writeheader()
        writer.writerows(new_rows)
    return len(new_rows)


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Match shrep Delivery City rows to rc Destination City rows."
    )
    parser.add_argument("input", nargs="?", default=Path("colector.xlsx"), type=Path)
    parser.add_argument(
        "-o",
        "--output",
        default=Path("outputs/city_matching.xlsx"),
        type=Path,
        help="Output .xlsx path.",
    )
    parser.add_argument(
        "--aliases",
        default=Path("city_aliases.csv"),
        type=Path,
        help="Optional approved alias CSV.",
    )
    parser.add_argument("--auto-threshold", default=90.0, type=float)
    parser.add_argument("--review-threshold", default=75.0, type=float)
    parser.add_argument("--min-margin", default=3.0, type=float)
    parser.add_argument(
        "--allow-cross-country",
        action="store_true",
        help="Allow matching outside Arrival Country when no same-country candidates exist.",
    )
    parser.add_argument(
        "--review",
        action="store_true",
        help="Open an interactive console review for uncertain city pairs.",
    )
    parser.add_argument(
        "--save-aliases",
        action="store_true",
        help="Append accepted/manual matches to the alias CSV.",
    )
    return parser.parse_args(argv)


def main() -> None:
    raw_args = sys.argv[1:]
    no_args_mode = len(raw_args) == 0
    frozen = bool(getattr(sys, "frozen", False))
    args = parse_args(raw_args)
    enable_utf8_console()
    print_banner()

    try:
        with step("Reading workbook"):
            rc, shrep = load_input(args.input)

        with step("Building candidate index"):
            candidates = build_candidates(rc)
            aliases = load_aliases(args.aliases)

        with step("Matching cities"):
            outputs = build_outputs(
                rc=rc,
                shrep=shrep,
                candidates=candidates,
                aliases=aliases,
                auto_threshold=args.auto_threshold,
                review_threshold=args.review_threshold,
                min_margin=args.min_margin,
                allow_cross_country=args.allow_cross_country,
            )

        review_rows = outputs["review_needed"]
        if review_rows and not args.review:
            print()
            print(f"Review available: {len(review_rows)} city pair(s) need a decision.")

        run_review = args.review
        if should_prompt_for_review(args.review, review_rows, no_args_mode):
            run_review = ask_yes_no("Run manual review now?", default=True)

        if run_review:
            outputs["city_map"] = interactive_review(outputs["city_map"])
            refresh_review_and_alias_sheets(
                outputs,
                shrep_rows=len(shrep),
                rc_rows=len(rc),
                auto_threshold=args.auto_threshold,
                review_threshold=args.review_threshold,
                min_margin=args.min_margin,
                allow_cross_country=args.allow_cross_country,
            )

        save_manual_aliases = args.save_aliases or (no_args_mode and run_review)
        if save_manual_aliases:
            manual_rows = [
                alias_row_from_match(row)
                for row in outputs["city_map"]
                if row.get("match_status") == "manual_matched"
            ]
            written = append_aliases(args.aliases, manual_rows)
            print(f"[ok] Saved {written} new alias row(s) to {args.aliases}")

        with step("Writing Excel output"):
            write_outputs(outputs, args.output)

        print_summary(outputs["summary"], outputs["city_map"], args.output)
    except Exception as exc:
        print()
        print(f"[error] {exc}")
        raise SystemExit(1) from exc
    finally:
        if should_pause_on_exit(no_args_mode, frozen):
            pause_console()


if __name__ == "__main__":
    main()
